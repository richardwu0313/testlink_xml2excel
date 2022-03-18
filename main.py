import os
from openpyxl import Workbook

import xmltodict
from loguru import logger


BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def get_testsuite_depth(dt):
    '''

    :param dt:
    :return: depth of testsuite
    '''
    depth = 0
    for item, value in dt.items():
        if item == 'testsuite':
            depth += 1
            if isinstance(value, dict):
                depth += get_testsuite_depth(dt=value)
            elif isinstance(value, list):
                sub_depth_list = []
                for sub_item in value:
                    # ignore deprecated folder
                    if sub_item.get('@name', '') == 'deprecated':
                        continue
                    sub_depth_list.append(get_testsuite_depth(sub_item))
                depth += max(sub_depth_list)

    return depth


def html2text(html):
    if html:
        return html.replace('&nbsp;', ' ').\
            replace('<p>', '').replace('</p>', '').\
            replace('<strong>', '').replace('</strong>', ''). \
            replace('<ul>', '').replace('</ul>', ''). \
            replace('<li>', '').replace('</li>', '')
    return ''


def to_testcase_data(testcase):
    # key, name, version, pre-conditions, keywords, step#, actions, expected_results, execution_type
    key = html2text(testcase.get('externalid', ''))
    name = html2text(testcase.get('@name', ''))
    version = html2text(testcase.get('version', ''))
    pre_conditions = html2text(testcase.get('preconditions', ''))
    keywords = []
    if 'keywords' in testcase:
        for kw, vl in testcase.get('keywords', []).items():
            if isinstance(vl, dict):
                keywords.append(html2text(vl.get('@name')))
            elif isinstance(vl, list):
                for item in vl:
                    keywords.append(html2text(item.get('@name')))
            else:
                logger.error(f'unexpected keywords: {kw} => {vl}')
    steps = []
    if 'steps' in testcase:
        if isinstance(testcase.get('steps').get('step'), dict):
            steps.append(
                (html2text(testcase.get('steps').get('step').get('step_number')),
                 html2text(testcase.get('steps').get('step').get('actions')),
                 html2text(testcase.get('steps').get('step').get('expectedresults')),
                 'Manual' if html2text(testcase.get('steps').get('step').get('execution_type')) == '1' else 'Automated',)
            )
        elif isinstance(testcase.get('steps').get('step'), list):
            for step in testcase.get('steps').get('step'):
                steps.append(
                    (html2text(step.get('step_number')),
                     html2text(step.get('actions')),
                     html2text(step.get('expectedresults')),
                     'Manual' if html2text(step.get('execution_type')) == '1' else 'Automated',)
                )
        else:
            logger.error(f'unexpected step: {testcase.get("steps").get("step")}')
    return {'key': key,
            'name': name,
            'version': version,
            'pre_conditions': pre_conditions,
            'keywords': keywords,
            'teststeps': steps}


def to_testsuite_data(testsuite):
    ts_key = testsuite.get('@id')
    ts_name = testsuite.get('@name')
    custom_field = testsuite.get('custom_fields').get('custom_field', '')
    ts_epic = ''
    if custom_field:
        ts_epic = custom_field.get('value')
    testcases = []
    if isinstance(testsuite.get('testcase', []), list):
        for item in testsuite.get('testcase', []):
            tc = to_testcase_data(item)
            testcases.append(tc)
    else:
        testcases.append(to_testcase_data(testsuite.get('testcase')))

    return {
        'key': ts_key,
        'name': ts_name,
        'epic': ts_epic,
        'testcases': testcases
    }


def to_testfeature_data(testfeature):
    tf_key = testfeature.get('@id')
    tf_name = testfeature.get('@name')
    testsuites = []
    if isinstance(testfeature.get('testsuite', []), list):
        for item in testfeature.get('testsuite', []):
            testsuites.append(to_testsuite_data(testsuite=item))
    else:
        testsuites.append(to_testsuite_data(testsuite=testfeature.get('testsuite')))
    return {
        'key': tf_key,
        'name': tf_name,
        'testsuites': testsuites
    }


def to_testproject_data(testproject):
    testfeatures = []
    for item in testproject.get('testsuite'):
        testfeatures.append(to_testfeature_data(testfeature=item))
    return {
        'testfeatures': testfeatures
    }


def generate_testcase_excel(data, name='testcase.xlsx'):
    wb = Workbook()
    ws = wb.create_sheet('testcase')
    ws.title = 'testlink'
    ws.append(['key', 'name', 'version', 'pre condition', 'key words',
               'step#', 'actions', 'expected result', 'execution type'])
    index = 1
    for step in data.get('teststeps'):
        ws.append(
            (data.get('key'),
             data.get('name'),
             data.get('version'),
             data.get('pre_conditions'),
             ''.join(data.get('keywords')),
             step[0],
             step[1],
             step[2],
             step[3],)
        )
    # merge test case info
    for cell in ('A', 'B', 'C', 'D', 'E'):
        ws.merge_cells(f'{cell}{index + 1}:{cell}{index + len(data.get("teststeps"))}')
    wb.save(filename=name)


def generate_testsuite_excel(data, name='testsuite.xlsx'):
    wb = Workbook()
    ws = wb.create_sheet('testsuite')
    ws.title = 'testlink'
    ws.append(['ts key', 'ts name', 'ts epic', 'key', 'name', 'version',
               'pre-conditions', 'keywords', 'step#', 'actions',
               'expected result', 'execution type'])
    index = 1
    for tc in data['testcases']:
        for step in tc['teststeps']:
            ws.append(
                (data.get('key'),
                 data.get('name'),
                 data.get('epic'),
                 tc.get('key'),
                 tc.get('name'),
                 tc.get('version'),
                 tc.get('pre_conditions'),
                 ' '.join(tc.get('keywords')),
                 step[0],
                 step[1],
                 step[2],
                 step[3],
                 )
            )
        for cell in ['D', 'E', 'F', 'G', 'H']:
            ws.merge_cells(f'{cell}{index + 1}:{cell}{index + len(tc["teststeps"])}')
        index += len(tc['teststeps'])
    for cell in ['A', 'B', 'C']:
        ws.merge_cells(f'{cell}2:{cell}{index}')
    wb.save(filename=name)


def generate_testfeature_excel(data, name='testfeature.xlsx'):
    wb = Workbook()
    ws = wb.create_sheet('testfeature')
    ws.title = 'testlink'
    ws.append(['feature name', 'feature key', 'ts key', 'ts name', 'ts epic', 'key', 'name', 'version',
               'pre-conditions', 'keywords', 'step#', 'actions',
               'expected result', 'execution type'])
    index = 1
    for tf in data['testsuites']:
        start = index
        for tc in tf['testcases']:
            for step in tc['teststeps']:
                ws.append(
                    (data.get('key'),
                     data.get('name'),
                     tf.get('key'),
                     tf.get('name'),
                     tf.get('epic'),
                     tc.get('key'),
                     tc.get('name'),
                     tc.get('version'),
                     tc.get('pre_conditions'),
                     ' '.join(tc.get('keywords')),
                     step[0],
                     step[1],
                     step[2],
                     step[3],
                     )
                )
            for cell in ['F', 'G', 'H', 'I', 'J']:
                ws.merge_cells(f'{cell}{index + 1}:{cell}{index + len(tc["teststeps"])}')
            index += len(tc['teststeps'])
        for cell in ['C', 'D', 'E']:
            ws.merge_cells(f'{cell}{start + 1}:{cell}{index}')
    for cell in ['A', 'B']:
        ws.merge_cells(f'{cell}2:{cell}{index}')
    wb.save(filename=name)


def generate_testproject_excel(data, name='testproject.xlsx'):
    wb = Workbook()
    ws = wb.create_sheet('testproject')
    ws.title = 'testlink'
    ws.append(['feature name', 'feature key', 'ts key', 'ts name', 'ts epic', 'key', 'name', 'version',
               'pre-conditions', 'keywords', 'step#', 'actions',
               'expected result', 'execution type'])
    index = 1
    for tp in data['testfeatures']:
        tp_start = index
        for tf in tp['testsuites']:
            ts_start = index
            for tc in tf['testcases']:
                for step in tc['teststeps']:
                    ws.append(
                        (data.get('key'),
                         data.get('name'),
                         tf.get('key'),
                         tf.get('name'),
                         tf.get('epic'),
                         tc.get('key'),
                         tc.get('name'),
                         tc.get('version'),
                         tc.get('pre_conditions'),
                         ' '.join(tc.get('keywords')),
                         step[0],
                         step[1],
                         step[2],
                         step[3],
                         )
                    )
                if len(tc["teststeps"]) > 0:
                    for cell in ['F', 'G', 'H', 'I', 'J']:
                        ws.merge_cells(f'{cell}{index + 1}:{cell}{index + len(tc["teststeps"])}')
                index += len(tc['teststeps'])
            if index > ts_start:
                # index == ts_start when empty test suite
                for cell in ['C', 'D', 'E']:
                    ws.merge_cells(f'{cell}{ts_start + 1}:{cell}{index}')
        if index > tp_start:
            # index == tp_start when empty test feature
            for cell in ['A', 'B']:
                ws.merge_cells(f'{cell}{tp_start + 1}:{cell}{index}')
    wb.save(filename=name)


def generate_excel(input, output='output.xlsx'):
    with open(input, 'r') as f:
        content = f.read()
    xd = xmltodict.parse(content)
    depth = get_testsuite_depth(dt=xd)
    if depth == 0:
        # /testcase
        tc_data = to_testcase_data(testcase=xd.get('testcases').get('testcase'))
        generate_testcase_excel(data=tc_data, name=output)
    elif depth == 1:
        # /testsuite/testcase
        ts_data = to_testsuite_data(testsuite=xd.get('testsuite'))
        generate_testsuite_excel(data=ts_data, name=output)
    elif depth == 2:
        # /testfeature/testsuite/testcase
        tf_data = to_testfeature_data(testfeature=xd.get('testsuite'))
        generate_testfeature_excel(data=tf_data, name=output)
    elif depth == 3:
        # /testproject/testfeature/testsuite/testcase
        tp_data = to_testproject_data(testproject=xd.get('testsuite'))
        generate_testproject_excel(data=tp_data)
    else:
        logger.error('invalid testcase structure')
        return False


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    from optparse import OptionParser

    parser = OptionParser()
    parser.add_option('-i', '--input_file', dest='input_file', help='input xml file name')
    parser.add_option('-o', '--output_file', dest='output_file', help='output excel file name')
    options, args = parser.parse_args()
    if not options.input_file:
        logger.error('no input xml file name')
        exit(-1)
    if not options.output_file:
        generate_excel(input=os.path.join(BASE_DIR, options.input_file))
    else:
        output_file = options.output_file
        generate_excel(input=os.path.join(BASE_DIR, options.input_file),
                       output=output_file)


